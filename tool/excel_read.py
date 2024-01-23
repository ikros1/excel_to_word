import openpyxl

def read_xlsx_file_and_return_replace_list(file_path):
    return_list = []

    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # First loop: Find columns with "er_wr_" and corresponding elements
    columns_with_marker = {}
    for col_idx in range(1, sheet.max_column + 1):
        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and "er_wr_" in str(cell_value):
                if col_idx not in columns_with_marker:
                    columns_with_marker[col_idx] = []
                columns_with_marker[col_idx].append((row_idx, cell_value))

    # Second loop: Build the return_list
    for row_idx in range(1, sheet.max_row + 1):
        row_data = []
        for col_idx, markers in columns_with_marker.items():
            if markers:
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    row_data.append((cell_value, markers[0][1]))
        if row_data:
            return_list.append(row_data)

    return return_list

# Example usage:
file_path = "excel_file/example.xlsx"
result = read_xlsx_file_and_return_replace_list(file_path)
