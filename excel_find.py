import openpyxl


def find_row_by_value(file_path, sheet_name, target_value):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    sheet = workbook[sheet_name]

    # 遍历每一行
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        # 在当前行中查找目标值
        if target_value in row:
            # 返回整行数据
            return row

    # 如果未找到目标值，返回None
    return None



file_path = 'excel_file/read_file.xlsx'
sheet_name = 'Sheet1'
target_value = '电容'

result_row = find_row_by_value(file_path, sheet_name, target_value)

if result_row:
    print("找到了目标值所在的行：", result_row)
else:
    print("未找到目标值。")
